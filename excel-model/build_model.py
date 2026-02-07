"""
Build the Personal Budget App Financial Model in Excel.
Based on the PRD (budget-app-prd.md).
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from copy import copy

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
DARK_BLUE = "1F3864"
MED_BLUE = "2E75B6"
LIGHT_BLUE = "D6E4F0"
GREEN = "548235"
LIGHT_GREEN = "E2EFDA"
RED = "C00000"
LIGHT_RED = "FCE4EC"
GRAY = "808080"
LIGHT_GRAY = "F2F2F2"
WHITE = "FFFFFF"
BLACK = "000000"

HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE, size=12)
HEADER_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
SUB_HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE, size=11)
SUB_HEADER_FILL = PatternFill(start_color=MED_BLUE, end_color=MED_BLUE, fill_type="solid")
CATEGORY_FONT = Font(name="Calibri", bold=True, size=11)
NORMAL_FONT = Font(name="Calibri", size=11)
MONEY_FORMAT = '#,##0.00'
PCT_FORMAT = '0.0%'
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def style_header_row(ws, row, max_col, font=HEADER_FONT, fill=HEADER_FILL):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def style_range(ws, min_row, max_row, min_col, max_col, font=NORMAL_FONT, num_fmt=None, fill=None):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = font
            cell.border = THIN_BORDER
            if num_fmt:
                cell.number_format = num_fmt
            if fill:
                cell.fill = fill


def auto_width(ws, min_col=1, max_col=None, extra=3):
    if max_col is None:
        max_col = ws.max_column
    for col in range(min_col, max_col + 1):
        max_len = 0
        letter = get_column_letter(col)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col, max_col=col):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + extra, 40)


# ===================================================================
# 1. REPORTING CATEGORIES SHEET
# ===================================================================
def build_categories_sheet(wb):
    ws = wb.create_sheet("Reporting Categories")

    # Title
    ws.merge_cells("A1:E1")
    ws["A1"] = "Personal Budget – Reporting Categories"
    ws["A1"].font = Font(name="Calibri", bold=True, color=WHITE, size=14)
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="center")
    for c in range(1, 6):
        ws.cell(row=1, column=c).fill = HEADER_FILL

    headers = ["#", "Category", "Type", "Description / Examples", "Plaid Mapping"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, 5, font=SUB_HEADER_FONT, fill=SUB_HEADER_FILL)

    categories = [
        (1, "Income", "Income",
         "Paycheck, transfers in, refunds, interest (positive)",
         "Income; Transfer > Payroll; Interest (positive)"),
        (2, "Housing", "Expense",
         "Rent, mortgage, property tax, home insurance",
         "Payment > Rent"),
        (3, "Bills & Utilities", "Expense",
         "Phone, internet, electric, water, gas, insurance premiums",
         "Utilities; Payment > *; Tax"),
        (4, "Groceries", "Expense",
         "Supermarkets, grocery stores, wholesale clubs",
         "Food and Drink > Groceries; Shops > Supermarkets"),
        (5, "Dining Out", "Expense",
         "Restaurants, fast food, coffee shops, bars, delivery",
         "Food and Drink > Restaurants; Food and Drink > *"),
        (6, "Transportation", "Expense",
         "Gas, parking, public transit, rideshare, car maintenance",
         "Travel > * (non-airline/lodging)"),
        (7, "Shopping", "Expense",
         "Retail, Amazon, clothing, electronics, general merchandise",
         "Shops > * (non-supermarket)"),
        (8, "Entertainment", "Expense",
         "Streaming services, movies, games, concerts, events",
         "Recreation"),
        (9, "Health", "Expense",
         "Pharmacy, doctor visits, gym memberships, medical bills",
         "Healthcare"),
        (10, "Travel", "Expense",
         "Flights, hotels, vacation rentals, travel booking",
         "Travel > Airlines; Travel > Lodging"),
        (11, "Subscriptions", "Expense",
         "Recurring services (SaaS, news, apps, memberships)",
         "Service > Subscription"),
        (12, "Personal Care", "Expense",
         "Haircuts, beauty, spa, cosmetics",
         "Service > Personal Care"),
        (13, "Education", "Expense",
         "Tuition, books, courses, school supplies",
         "Education-related services"),
        (14, "Gifts & Donations", "Expense",
         "Charitable donations, gifts for others",
         "Gifts / Donations categories"),
        (15, "Fees & Charges", "Expense",
         "Bank fees, ATM fees, late fees, overdraft charges",
         "Bank Fees; Interest (negative)"),
        (16, "Transfer", "Neutral",
         "Internal transfers, Venmo/Zelle (often noise, excluded from totals)",
         "Transfer > Internal; Transfer > *"),
        (17, "Other / Uncategorized", "Expense",
         "Anything that doesn't fit above; fallback category",
         "Cash Advance; Community; Service > *"),
    ]

    for idx, (num, cat, typ, desc, plaid) in enumerate(categories):
        r = 4 + idx
        ws.cell(row=r, column=1, value=num)
        ws.cell(row=r, column=2, value=cat)
        ws.cell(row=r, column=3, value=typ)
        ws.cell(row=r, column=4, value=desc)
        ws.cell(row=r, column=5, value=plaid)

        # Color-code type
        if typ == "Income":
            fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
        elif typ == "Neutral":
            fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
        else:
            fill = None

        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(c >= 4))
            if fill:
                cell.fill = fill

        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=2).font = CATEGORY_FONT
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="center", vertical="center")

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 45

    return ws


# ===================================================================
# 2. PLAID CATEGORY MAPPING SHEET
# ===================================================================
def build_plaid_mapping_sheet(wb):
    ws = wb.create_sheet("Plaid Category Mapping")

    ws.merge_cells("A1:C1")
    ws["A1"] = "Plaid Category → App Category Mapping"
    ws["A1"].font = Font(name="Calibri", bold=True, color=WHITE, size=14)
    ws["A1"].fill = HEADER_FILL
    for c in range(1, 4):
        ws.cell(row=1, column=c).fill = HEADER_FILL

    headers = ["Plaid Category (Primary > Secondary)", "App Category", "Notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, 3, font=SUB_HEADER_FONT, fill=SUB_HEADER_FILL)

    mappings = [
        ("Income", "Income", "Direct mapping"),
        ("Transfer > Payroll", "Income", "Payroll deposits"),
        ("Interest (positive)", "Income", "Savings interest, dividends"),
        ("Bank Fees", "Fees & Charges", "All bank fee types"),
        ("Interest (negative)", "Fees & Charges", "Interest charges on credit"),
        ("Cash Advance", "Other", "Rare; flag for review"),
        ("Community", "Other", "Community-related services"),
        ("Food and Drink > Restaurants", "Dining Out", "Includes fast food, cafes"),
        ("Food and Drink > Groceries", "Groceries", "Grocery-specific merchants"),
        ("Food and Drink > Coffee Shop", "Dining Out", "Starbucks, etc."),
        ("Food and Drink > *", "Dining Out", "Catch-all food & drink"),
        ("Healthcare", "Health", "All healthcare sub-categories"),
        ("Payment > Rent", "Housing", "Rent payments"),
        ("Payment > Mortgage", "Housing", "Mortgage payments"),
        ("Payment > *", "Bills & Utilities", "Other bill payments"),
        ("Recreation", "Entertainment", "All recreation sub-categories"),
        ("Service > Subscription", "Subscriptions", "Recurring service charges"),
        ("Service > Personal Care", "Personal Care", "Haircuts, beauty, spa"),
        ("Service > Financial", "Fees & Charges", "Financial service fees"),
        ("Service > *", "Other", "Catch-all services"),
        ("Shops > Supermarkets", "Groceries", "Supermarket/wholesale"),
        ("Shops > Clothing", "Shopping", "Apparel stores"),
        ("Shops > Electronics", "Shopping", "Tech retail"),
        ("Shops > *", "Shopping", "Catch-all retail"),
        ("Tax", "Bills & Utilities", "Tax payments"),
        ("Transfer > Internal", "Transfer", "Account-to-account; exclude from totals"),
        ("Transfer > Wire", "Transfer", "Wire transfers"),
        ("Transfer > *", "Transfer", "Catch-all transfers"),
        ("Travel > Airlines", "Travel", "Flight bookings"),
        ("Travel > Lodging", "Travel", "Hotels, vacation rentals"),
        ("Travel > Car Rental", "Transportation", "Rental cars"),
        ("Travel > Taxi / Rideshare", "Transportation", "Uber, Lyft, taxis"),
        ("Travel > Gas", "Transportation", "Gas stations"),
        ("Travel > Parking", "Transportation", "Parking fees"),
        ("Travel > Public Transit", "Transportation", "Buses, trains, metro"),
        ("Travel > *", "Transportation", "Other travel"),
        ("Utilities", "Bills & Utilities", "Electric, water, gas"),
    ]

    for idx, (plaid, app, notes) in enumerate(mappings):
        r = 4 + idx
        ws.cell(row=r, column=1, value=plaid)
        ws.cell(row=r, column=2, value=app)
        ws.cell(row=r, column=3, value=notes)
        for c in range(1, 4):
            cell = ws.cell(row=r, column=c)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 40

    return ws


# ===================================================================
# 3. MONTHLY BUDGET TEMPLATE SHEET
# ===================================================================
def build_monthly_budget_sheet(wb):
    ws = wb.create_sheet("Monthly Budget")

    ws.merge_cells("A1:N1")
    ws["A1"] = "Monthly Budget – 12-Month Plan"
    ws["A1"].font = Font(name="Calibri", bold=True, color=WHITE, size=14)
    ws["A1"].fill = HEADER_FILL
    for c in range(1, 15):
        ws.cell(row=1, column=c).fill = HEADER_FILL

    # Headers: Category | Jan | Feb | ... | Dec | Annual Total
    headers = ["Category"] + MONTHS + ["Annual Total"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, 14, font=SUB_HEADER_FONT, fill=SUB_HEADER_FILL)

    # ---- INCOME SECTION ----
    income_fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
    r = 4
    ws.cell(row=r, column=1, value="INCOME")
    ws.cell(row=r, column=1).font = Font(name="Calibri", bold=True, size=11, color=GREEN)
    for c in range(1, 15):
        ws.cell(row=r, column=c).fill = income_fill
        ws.cell(row=r, column=c).border = THIN_BORDER

    income_items = ["Salary / Paycheck", "Side Income", "Interest / Dividends", "Refunds", "Other Income"]
    for idx, item in enumerate(income_items):
        r = 5 + idx
        ws.cell(row=r, column=1, value=item).font = NORMAL_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        for c in range(2, 14):
            cell = ws.cell(row=r, column=c, value=0)
            cell.number_format = MONEY_FORMAT
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
        # Annual total formula
        cell = ws.cell(row=r, column=14)
        cell.value = f"=SUM(B{r}:M{r})"
        cell.number_format = MONEY_FORMAT
        cell.font = Font(name="Calibri", bold=True, size=11)
        cell.border = THIN_BORDER

    # Total Income row
    r_total_income = 5 + len(income_items)
    ws.cell(row=r_total_income, column=1, value="TOTAL INCOME")
    ws.cell(row=r_total_income, column=1).font = Font(name="Calibri", bold=True, size=11, color=GREEN)
    for c in range(1, 15):
        ws.cell(row=r_total_income, column=c).fill = income_fill
        ws.cell(row=r_total_income, column=c).border = THIN_BORDER
        ws.cell(row=r_total_income, column=c).font = Font(name="Calibri", bold=True, size=11, color=GREEN)
    for c in range(2, 15):
        col_letter = get_column_letter(c)
        ws.cell(row=r_total_income, column=c).value = f"=SUM({col_letter}5:{col_letter}{r_total_income-1})"
        ws.cell(row=r_total_income, column=c).number_format = MONEY_FORMAT

    # ---- EXPENSE SECTION ----
    r_exp_header = r_total_income + 2
    expense_fill_header = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")
    ws.cell(row=r_exp_header, column=1, value="EXPENSES")
    ws.cell(row=r_exp_header, column=1).font = Font(name="Calibri", bold=True, size=11, color=RED)
    for c in range(1, 15):
        ws.cell(row=r_exp_header, column=c).fill = expense_fill_header
        ws.cell(row=r_exp_header, column=c).border = THIN_BORDER

    expense_categories = [
        ("Housing", ["Rent / Mortgage", "Property Tax", "Home Insurance", "HOA Fees", "Home Maintenance"]),
        ("Bills & Utilities", ["Electric", "Water / Sewer", "Gas (home)", "Internet", "Phone", "Insurance (auto/health/life)"]),
        ("Groceries", ["Grocery Stores", "Wholesale Clubs (Costco, etc.)"]),
        ("Dining Out", ["Restaurants", "Fast Food", "Coffee Shops", "Food Delivery"]),
        ("Transportation", ["Gas (auto)", "Car Payment", "Car Insurance", "Parking", "Public Transit", "Rideshare (Uber/Lyft)", "Car Maintenance"]),
        ("Shopping", ["Clothing", "Electronics", "Amazon / Online", "Home Goods", "General Retail"]),
        ("Entertainment", ["Streaming Services", "Movies / Events", "Games", "Hobbies"]),
        ("Health", ["Pharmacy", "Doctor / Medical", "Gym Membership", "Dental / Vision"]),
        ("Travel", ["Flights", "Hotels / Lodging", "Vacation Activities"]),
        ("Subscriptions", ["Software / SaaS", "News / Media", "App Subscriptions"]),
        ("Personal Care", ["Haircuts / Salon", "Beauty / Cosmetics"]),
        ("Education", ["Tuition / Courses", "Books / Supplies"]),
        ("Gifts & Donations", ["Charitable Donations", "Gifts"]),
        ("Fees & Charges", ["Bank Fees", "ATM Fees", "Late Fees / Penalties"]),
        ("Other / Uncategorized", ["Miscellaneous"]),
    ]

    current_row = r_exp_header + 1
    category_total_rows = []
    alt_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")

    for cat_idx, (cat_name, sub_items) in enumerate(expense_categories):
        # Category header
        ws.cell(row=current_row, column=1, value=cat_name)
        ws.cell(row=current_row, column=1).font = CATEGORY_FONT
        cat_fill = alt_fill if cat_idx % 2 == 0 else None
        for c in range(1, 15):
            ws.cell(row=current_row, column=c).border = THIN_BORDER
            if cat_fill:
                ws.cell(row=current_row, column=c).fill = cat_fill

        first_sub_row = current_row + 1
        for sub_idx, sub in enumerate(sub_items):
            sr = current_row + 1 + sub_idx
            ws.cell(row=sr, column=1, value=f"    {sub}").font = NORMAL_FONT
            ws.cell(row=sr, column=1).border = THIN_BORDER
            for c in range(2, 14):
                cell = ws.cell(row=sr, column=c, value=0)
                cell.number_format = MONEY_FORMAT
                cell.font = NORMAL_FONT
                cell.border = THIN_BORDER
            # Annual total
            cell = ws.cell(row=sr, column=14)
            cell.value = f"=SUM(B{sr}:M{sr})"
            cell.number_format = MONEY_FORMAT
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
        last_sub_row = current_row + len(sub_items)

        # Category subtotal (in the category header row cols 2-14)
        for c in range(2, 15):
            col_letter = get_column_letter(c)
            ws.cell(row=current_row, column=c).value = f"=SUM({col_letter}{first_sub_row}:{col_letter}{last_sub_row})"
            ws.cell(row=current_row, column=c).number_format = MONEY_FORMAT
            ws.cell(row=current_row, column=c).font = CATEGORY_FONT

        category_total_rows.append(current_row)
        current_row = last_sub_row + 1

    # Total Expenses row
    r_total_exp = current_row
    ws.cell(row=r_total_exp, column=1, value="TOTAL EXPENSES")
    ws.cell(row=r_total_exp, column=1).font = Font(name="Calibri", bold=True, size=11, color=RED)
    for c in range(1, 15):
        ws.cell(row=r_total_exp, column=c).fill = expense_fill_header
        ws.cell(row=r_total_exp, column=c).border = THIN_BORDER
        ws.cell(row=r_total_exp, column=c).font = Font(name="Calibri", bold=True, size=11, color=RED)
    for c in range(2, 15):
        col_letter = get_column_letter(c)
        refs = "+".join(f"{col_letter}{row}" for row in category_total_rows)
        ws.cell(row=r_total_exp, column=c).value = f"={refs}"
        ws.cell(row=r_total_exp, column=c).number_format = MONEY_FORMAT

    # ---- NET CASH FLOW ----
    r_net = r_total_exp + 2
    ws.cell(row=r_net, column=1, value="NET CASH FLOW")
    ws.cell(row=r_net, column=1).font = Font(name="Calibri", bold=True, size=12)
    for c in range(1, 15):
        ws.cell(row=r_net, column=c).border = THIN_BORDER
        ws.cell(row=r_net, column=c).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        ws.cell(row=r_net, column=c).font = Font(name="Calibri", bold=True, size=11)
    for c in range(2, 15):
        col_letter = get_column_letter(c)
        ws.cell(row=r_net, column=c).value = f"={col_letter}{r_total_income}-{col_letter}{r_total_exp}"
        ws.cell(row=r_net, column=c).number_format = MONEY_FORMAT

    # ---- SAVINGS RATE ----
    r_sav = r_net + 1
    ws.cell(row=r_sav, column=1, value="SAVINGS RATE")
    ws.cell(row=r_sav, column=1).font = Font(name="Calibri", bold=True, size=11)
    for c in range(1, 15):
        ws.cell(row=r_sav, column=c).border = THIN_BORDER
        ws.cell(row=r_sav, column=c).font = Font(name="Calibri", bold=True, size=11)
    for c in range(2, 15):
        col_letter = get_column_letter(c)
        ws.cell(row=r_sav, column=c).value = f"=IFERROR({col_letter}{r_net}/{col_letter}{r_total_income},0)"
        ws.cell(row=r_sav, column=c).number_format = PCT_FORMAT

    # Column widths
    ws.column_dimensions["A"].width = 30
    for c in range(2, 15):
        ws.column_dimensions[get_column_letter(c)].width = 14

    # Freeze panes
    ws.freeze_panes = "B4"

    return ws, r_total_income, r_total_exp, r_net, category_total_rows


# ===================================================================
# 4. DASHBOARD SHEET
# ===================================================================
def build_dashboard_sheet(wb, r_total_income, r_total_exp, r_net):
    ws = wb.create_sheet("Dashboard")

    ws.merge_cells("A1:H1")
    ws["A1"] = "Budget Dashboard – Monthly Summary"
    ws["A1"].font = Font(name="Calibri", bold=True, color=WHITE, size=14)
    ws["A1"].fill = HEADER_FILL
    for c in range(1, 9):
        ws.cell(row=1, column=c).fill = HEADER_FILL

    # Instructions
    ws.merge_cells("A3:H3")
    ws["A3"] = "This dashboard references the 'Monthly Budget' sheet. Enter your data there and this will update automatically."
    ws["A3"].font = Font(name="Calibri", italic=True, color=GRAY, size=10)

    # Annual summary cards
    r = 5
    cards = [
        ("Total Annual Income", f"='Monthly Budget'!N{r_total_income}", GREEN, LIGHT_GREEN),
        ("Total Annual Expenses", f"='Monthly Budget'!N{r_total_exp}", RED, LIGHT_RED),
        ("Annual Net Cash Flow", f"='Monthly Budget'!N{r_net}", DARK_BLUE, LIGHT_BLUE),
    ]

    for idx, (label, formula, color, bg_color) in enumerate(cards):
        col_start = 1 + idx * 3
        col_end = col_start + 1
        ws.merge_cells(start_row=r, start_column=col_start, end_row=r, end_column=col_end)
        cell = ws.cell(row=r, column=col_start, value=label)
        cell.font = Font(name="Calibri", bold=True, color=WHITE, size=11)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=col_end).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        ws.merge_cells(start_row=r+1, start_column=col_start, end_row=r+1, end_column=col_end)
        val_cell = ws.cell(row=r+1, column=col_start, value=formula)
        val_cell.font = Font(name="Calibri", bold=True, size=16)
        val_cell.number_format = MONEY_FORMAT
        val_cell.alignment = Alignment(horizontal="center")
        val_cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        ws.cell(row=r+1, column=col_end).fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")

    # Monthly summary table
    r = 9
    headers = ["Metric"] + MONTHS
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_header_row(ws, r, 13, font=SUB_HEADER_FONT, fill=SUB_HEADER_FILL)

    metrics = [
        ("Total Income", r_total_income, GREEN),
        ("Total Expenses", r_total_exp, RED),
        ("Net Cash Flow", r_net, DARK_BLUE),
    ]

    for m_idx, (label, src_row, color) in enumerate(metrics):
        mr = r + 1 + m_idx
        ws.cell(row=mr, column=1, value=label).font = Font(name="Calibri", bold=True, color=color, size=11)
        ws.cell(row=mr, column=1).border = THIN_BORDER
        for c in range(2, 14):
            col_letter = get_column_letter(c)
            ws.cell(row=mr, column=c).value = f"='Monthly Budget'!{col_letter}{src_row}"
            ws.cell(row=mr, column=c).number_format = MONEY_FORMAT
            ws.cell(row=mr, column=c).font = NORMAL_FONT
            ws.cell(row=mr, column=c).border = THIN_BORDER

    # Column widths
    ws.column_dimensions["A"].width = 20
    for c in range(2, 14):
        ws.column_dimensions[get_column_letter(c)].width = 13

    return ws


# ===================================================================
# 5. CATEGORY BUDGET vs ACTUAL SHEET
# ===================================================================
def build_budget_vs_actual_sheet(wb):
    ws = wb.create_sheet("Budget vs Actual")

    ws.merge_cells("A1:F1")
    ws["A1"] = "Budget vs Actual – Monthly Comparison"
    ws["A1"].font = Font(name="Calibri", bold=True, color=WHITE, size=14)
    ws["A1"].fill = HEADER_FILL
    for c in range(1, 7):
        ws.cell(row=1, column=c).fill = HEADER_FILL

    ws.cell(row=3, column=1, value="Select Month:")
    ws.cell(row=3, column=1).font = CATEGORY_FONT
    ws.cell(row=3, column=2, value="January")
    ws.cell(row=3, column=2).font = Font(name="Calibri", bold=True, size=11, color=MED_BLUE)

    headers = ["Category", "Budget", "Actual", "Variance ($)", "Variance (%)", "Status"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=5, column=i, value=h)
    style_header_row(ws, 5, 6, font=SUB_HEADER_FONT, fill=SUB_HEADER_FILL)

    expense_cats = [
        "Housing", "Bills & Utilities", "Groceries", "Dining Out",
        "Transportation", "Shopping", "Entertainment", "Health",
        "Travel", "Subscriptions", "Personal Care", "Education",
        "Gifts & Donations", "Fees & Charges", "Other / Uncategorized"
    ]

    for idx, cat in enumerate(expense_cats):
        r = 6 + idx
        ws.cell(row=r, column=1, value=cat).font = NORMAL_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        # Budget column (to be filled in)
        cell_b = ws.cell(row=r, column=2, value=0)
        cell_b.number_format = MONEY_FORMAT
        cell_b.border = THIN_BORDER
        # Actual column (to be filled in)
        cell_a = ws.cell(row=r, column=3, value=0)
        cell_a.number_format = MONEY_FORMAT
        cell_a.border = THIN_BORDER
        # Variance $
        cell_v = ws.cell(row=r, column=4)
        cell_v.value = f"=B{r}-C{r}"
        cell_v.number_format = MONEY_FORMAT
        cell_v.border = THIN_BORDER
        # Variance %
        cell_vp = ws.cell(row=r, column=5)
        cell_vp.value = f"=IFERROR((B{r}-C{r})/B{r},0)"
        cell_vp.number_format = PCT_FORMAT
        cell_vp.border = THIN_BORDER
        # Status
        cell_s = ws.cell(row=r, column=6)
        cell_s.value = f'=IF(C{r}<=B{r},"Under Budget","Over Budget")'
        cell_s.border = THIN_BORDER

    # Totals row
    r_tot = 6 + len(expense_cats)
    ws.cell(row=r_tot, column=1, value="TOTAL").font = Font(name="Calibri", bold=True, size=11)
    for c in range(1, 7):
        ws.cell(row=r_tot, column=c).border = THIN_BORDER
        ws.cell(row=r_tot, column=c).font = Font(name="Calibri", bold=True, size=11)
        ws.cell(row=r_tot, column=c).fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
    for c in [2, 3, 4]:
        col_letter = get_column_letter(c)
        ws.cell(row=r_tot, column=c).value = f"=SUM({col_letter}6:{col_letter}{r_tot-1})"
        ws.cell(row=r_tot, column=c).number_format = MONEY_FORMAT
    ws.cell(row=r_tot, column=5).value = f"=IFERROR((B{r_tot}-C{r_tot})/B{r_tot},0)"
    ws.cell(row=r_tot, column=5).number_format = PCT_FORMAT
    ws.cell(row=r_tot, column=6).value = f'=IF(C{r_tot}<=B{r_tot},"Under Budget","Over Budget")'

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 16

    return ws


# ===================================================================
# 6. TRANSACTION LOG TEMPLATE
# ===================================================================
def build_transaction_log_sheet(wb):
    ws = wb.create_sheet("Transaction Log")

    ws.merge_cells("A1:H1")
    ws["A1"] = "Transaction Log – Raw Transaction Data"
    ws["A1"].font = Font(name="Calibri", bold=True, color=WHITE, size=14)
    ws["A1"].fill = HEADER_FILL
    for c in range(1, 9):
        ws.cell(row=1, column=c).fill = HEADER_FILL

    headers = ["Date", "Merchant / Description", "Amount", "Plaid Category",
               "App Category", "Account", "Pending", "Notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, 8, font=SUB_HEADER_FONT, fill=SUB_HEADER_FILL)

    # Sample data rows
    samples = [
        ("2026-01-15", "Whole Foods Market", 87.32, "Food and Drink > Groceries", "Groceries", "Checking ****1234", "No", ""),
        ("2026-01-15", "Uber Eats", 32.50, "Food and Drink > Restaurants > Fast Food", "Dining Out", "Credit Card ****5678", "No", ""),
        ("2026-01-14", "Shell Gas Station", 45.00, "Travel > Gas", "Transportation", "Credit Card ****5678", "No", ""),
        ("2026-01-14", "Netflix", 15.99, "Service > Subscription", "Subscriptions", "Credit Card ****5678", "No", "Monthly"),
        ("2026-01-13", "Target", 124.56, "Shops > General Merchandise", "Shopping", "Credit Card ****5678", "Yes", "Pending"),
        ("2026-01-13", "Employer Direct Deposit", -3500.00, "Transfer > Payroll", "Income", "Checking ****1234", "No", "Bi-weekly pay"),
        ("2026-01-12", "Walgreens Pharmacy", 22.45, "Healthcare > Pharmacy", "Health", "Credit Card ****5678", "No", ""),
        ("2026-01-12", "Spotify", 9.99, "Service > Subscription", "Subscriptions", "Credit Card ****5678", "No", "Monthly"),
        ("2026-01-11", "Rent Payment", 1800.00, "Payment > Rent", "Housing", "Checking ****1234", "No", "Monthly rent"),
        ("2026-01-10", "Venmo Transfer", 50.00, "Transfer > Third Party", "Transfer", "Checking ****1234", "No", "Split dinner"),
    ]

    for idx, (date, merchant, amount, plaid_cat, app_cat, account, pending, notes) in enumerate(samples):
        r = 4 + idx
        ws.cell(row=r, column=1, value=date).font = NORMAL_FONT
        ws.cell(row=r, column=2, value=merchant).font = NORMAL_FONT
        cell_amt = ws.cell(row=r, column=3, value=amount)
        cell_amt.number_format = MONEY_FORMAT
        cell_amt.font = NORMAL_FONT
        if amount < 0:
            cell_amt.font = Font(name="Calibri", size=11, color=GREEN)
        ws.cell(row=r, column=4, value=plaid_cat).font = NORMAL_FONT
        ws.cell(row=r, column=5, value=app_cat).font = NORMAL_FONT
        ws.cell(row=r, column=6, value=account).font = NORMAL_FONT
        ws.cell(row=r, column=7, value=pending).font = NORMAL_FONT
        ws.cell(row=r, column=8, value=notes).font = NORMAL_FONT
        for c in range(1, 9):
            ws.cell(row=r, column=c).border = THIN_BORDER

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 38
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 20

    # Freeze panes
    ws.freeze_panes = "A4"

    return ws


# ===================================================================
# 7. ASSUMPTIONS & NOTES SHEET
# ===================================================================
def build_assumptions_sheet(wb):
    ws = wb.create_sheet("Assumptions & Notes")

    ws.merge_cells("A1:C1")
    ws["A1"] = "Model Assumptions & Notes"
    ws["A1"].font = Font(name="Calibri", bold=True, color=WHITE, size=14)
    ws["A1"].fill = HEADER_FILL
    for c in range(1, 4):
        ws.cell(row=1, column=c).fill = HEADER_FILL

    notes = [
        ("Data Source", "Plaid API (Sandbox mode for Phase 1)", "Production Plaid requires OAuth approval"),
        ("Category System", "17 categories mapped from Plaid's 3-level hierarchy", "User overrides take priority over Plaid auto-categorization"),
        ("Transfer Handling", "Transfers excluded from income/expense totals", "Categorized separately to avoid double-counting"),
        ("Pending Transactions", "Included in log with 'Pending' flag", "Will update on next Plaid sync"),
        ("Income Recognition", "Income = negative amounts in Plaid (money flowing in)", "Payroll, refunds, interest (positive)"),
        ("Expense Recognition", "Expenses = positive amounts in Plaid (money flowing out)", "All spending categories"),
        ("Currency", "USD only for Phase 1", "Multi-currency support out of scope"),
        ("Time Period", "Monthly reporting periods", "Calendar month boundaries"),
        ("Budget Method", "Zero-based budgeting approach", "Every dollar assigned a category"),
        ("Savings Rate", "= Net Cash Flow / Total Income", "Target: 20%+ for healthy finances"),
    ]

    headers = ["Topic", "Assumption", "Notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, 3, font=SUB_HEADER_FONT, fill=SUB_HEADER_FILL)

    for idx, (topic, assumption, note) in enumerate(notes):
        r = 4 + idx
        ws.cell(row=r, column=1, value=topic).font = CATEGORY_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=2, value=assumption).font = NORMAL_FONT
        ws.cell(row=r, column=2).border = THIN_BORDER
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
        ws.cell(row=r, column=3, value=note).font = Font(name="Calibri", italic=True, color=GRAY, size=11)
        ws.cell(row=r, column=3).border = THIN_BORDER
        ws.cell(row=r, column=3).alignment = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 50

    return ws


# ===================================================================
# MAIN: Build the workbook
# ===================================================================
def main():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Build all sheets
    build_categories_sheet(wb)
    build_plaid_mapping_sheet(wb)
    ws_budget, r_income, r_exp, r_net, cat_rows = build_monthly_budget_sheet(wb)
    build_dashboard_sheet(wb, r_income, r_exp, r_net)
    build_budget_vs_actual_sheet(wb)
    build_transaction_log_sheet(wb)
    build_assumptions_sheet(wb)

    # Set Dashboard as the active sheet
    wb.active = wb.sheetnames.index("Dashboard")

    output_path = "/home/user/themightyscot23/excel-model/personal-budget-financial-model.xlsx"
    wb.save(output_path)
    print(f"Financial model saved to: {output_path}")
    print(f"Sheets created: {wb.sheetnames}")


if __name__ == "__main__":
    main()
